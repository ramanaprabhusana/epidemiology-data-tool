#!/usr/bin/env python3
"""
Refresh Workbook: patches CLL_Consolidated_Forecast_Model_v5.2.xlsx in-place with fresh
pipeline outputs.

Touches ONLY the sheets the pipeline owns: never clobbers user edits on Forecast sheets:
  - Lookup Tables → Model Inputs (rows 5-10): refreshes incidence_rate, prevalence_anchor,
    mortality_anchor, mortality_apc per indication IF the pipeline delivers newer values.
  - <WorkbookKey> Evidence (e.g. Hodgkin Evidence): replaces evidence records wholesale with
    the latest pipeline export. Pipeline CLI uses long names (e.g. Hodgkin Lymphoma); those
    map to short workbook keys (Hodgkin, NonHodgkin, …).
  - Pipeline Metadata: stamps last-run timestamp, user, per-source row counts, checksums.

Forecast, Summary Dashboard, Population Projections, Methodology, Control Panel,
EPI Sources, Scenarios, Sensitivity: all untouched.

Usage:
    # Called automatically from run_all_indications.py
    # Or standalone:
    python refresh_workbook.py \\
        --workbook "../Client presentation/CLL_Consolidated_Forecast_Model_v5.2.xlsx" \\
        --manifest output/run_manifest.json
"""

from __future__ import annotations

import argparse
import json
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill

# Row mapping must match generate_multi_indication_workbook.py::MODEL_INPUT_ROWS (SSOT there).
MODEL_INPUT_ROWS = {
    "CLL":        5,
    "Hodgkin":    6,
    "NonHodgkin": 7,
    "Gastric":    8,
    "Ovarian":    9,
    "Prostate":   10,
}

# run_tool / run_all_indications indication strings → workbook sheet / Lookup row keys
PIPELINE_TO_WORKBOOK_KEY: Dict[str, str] = {
    "CLL": "CLL",
    "Hodgkin Lymphoma": "Hodgkin",
    "Hodgkin": "Hodgkin",
    "Non-Hodgkin Lymphoma": "NonHodgkin",
    "Non-Hodgkin": "NonHodgkin",
    "NHL": "NonHodgkin",
    "Gastric": "Gastric",
    "Ovarian": "Ovarian",
    "Prostate": "Prostate",
}


def _safe_name(s: str) -> str:
    """Match src.pipeline.runner._safe_name for output filenames."""
    return s.replace(" ", "_").replace(",", "").strip() or "unknown"


def _file_suffix(indication: str, country: Optional[str]) -> str:
    """Match runner file_suffix (evidence_by_metric_*, kpi_scorecard_*)."""
    ind_safe = _safe_name(indication)
    country_safe = _safe_name(country) if country else ""
    file_suffix = f"{ind_safe}_{country_safe}" if country_safe else ind_safe
    if file_suffix.endswith("_"):
        file_suffix = file_suffix[:-1]
    return file_suffix


def _workbook_zip_has_vba(workbook_path: Path) -> bool:
    """True only if the package contains a VBA project (macro-enabled)."""
    try:
        with zipfile.ZipFile(workbook_path, "r") as zf:
            return any(
                n.startswith("xl/vba") or n == "xl/vbaProject.bin"
                for n in zf.namelist()
            )
    except (OSError, zipfile.BadZipFile):
        return False


def pipeline_to_workbook_key(indication: str) -> Optional[str]:
    """Map pipeline indication label to v5.2 workbook keys (Lookup + Evidence sheet prefix)."""
    ind = (indication or "").strip()
    if not ind:
        return None
    if ind in MODEL_INPUT_ROWS:
        return ind
    return PIPELINE_TO_WORKBOOK_KEY.get(ind)

# Column mapping must match generate_multi_indication_workbook.py::MODEL_INPUT_COLS
# (column letter → parameter name for easy lookup when patching)
COL_TO_PARAM = {
    "B": "inc_rate",        # incidence rate per 100k
    "C": "prev_anchor",     # prevalence anchor 2022
    "D": "mort_anchor",     # mortality anchor 2022
    "E": "mort_apc",        # mortality APC
    # F-L are cascade ratios; we leave those alone (curated from literature, not auto-refreshed)
}


def _find_metric_from_df(df: pd.DataFrame, metric_ids: List[str]) -> Optional[float]:
    """Scan the pipeline's KPI scorecard / evidence DataFrame for the first matching metric.

    Returns the numeric 'value' column (attempts float conversion) or None if no match.
    """
    if df is None or df.empty:
        return None
    candidate_cols = [c for c in ("metric_id", "metric", "id") if c in df.columns]
    value_col = next((c for c in ("value", "value_num", "numeric_value") if c in df.columns), None)
    if not candidate_cols or not value_col:
        return None
    for col in candidate_cols:
        for mid in metric_ids:
            hits = df[df[col].astype(str).str.lower() == mid.lower()]
            if not hits.empty:
                raw = hits.iloc[0][value_col]
                try:
                    return float(raw)
                except (TypeError, ValueError):
                    continue
    return None


def _load_kpi_scorecard(scorecard_path: Path) -> Optional[pd.DataFrame]:
    if not scorecard_path.exists():
        return None
    try:
        return pd.read_csv(scorecard_path, encoding="utf-8")
    except Exception:
        return None


def patch_model_inputs(ws, indication: str, scorecard: pd.DataFrame) -> Dict[str, Any]:
    """Patch the 4 core parameters (inc_rate, prev_anchor, mort_anchor, mort_apc).
    Returns a dict summarizing what changed."""
    changed: Dict[str, Any] = {}
    row = MODEL_INPUT_ROWS.get(indication)
    if row is None:
        return changed

    pulls = {
        "B": _find_metric_from_df(scorecard, ["incidence_rate", "incidence_rate_crude", "incidence_rate_age_adjusted"]),
        "C": _find_metric_from_df(scorecard, ["prevalence"]),
        "D": _find_metric_from_df(scorecard, ["mortality", "mortality_deaths"]),
        # Mortality APC typically not in the evidence table; skip auto-patching unless we add it
    }
    for col, new_val in pulls.items():
        if new_val is None:
            continue
        cell = ws[f"{col}{row}"]
        old_val = cell.value
        # Only update when meaningfully different
        try:
            if old_val is None or abs(float(old_val) - float(new_val)) / max(abs(float(old_val)), 1e-6) > 0.001:
                cell.value = new_val
                # flag the cell with a light-yellow fill to show it was auto-updated
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                changed[COL_TO_PARAM[col]] = {"from": old_val, "to": new_val}
        except (TypeError, ValueError):
            pass
    return changed


def replace_evidence_sheet(wb, indication: str, evidence_df: pd.DataFrame) -> int:
    """Clear and re-populate <Indication> Evidence sheet. Returns rows written."""
    sheet_name = f"{indication} Evidence"
    if sheet_name not in wb.sheetnames:
        return 0
    ws = wb[sheet_name]
    # Don't touch rows 1-3 (title, subtitle, header). Clear rows 4 downward.
    max_row = ws.max_row
    if max_row >= 4:
        ws.delete_rows(4, max_row - 3)

    if evidence_df is None or evidence_df.empty:
        ws.cell(row=2, column=1, value=f"0 records  |  Refreshed: {datetime.now():%Y-%m-%d %H:%M}  |  Pipeline returned no rows.")
        return 0

    # Map pipeline columns to workbook's 13-column schema
    expected_cols = ["Category", "Metric", "Value", "Unit", "Source", "Source Tier",
                     "Confidence", "Confidence Score", "Year", "Geography",
                     "Population", "Definition", "Notes"]
    colmap = {
        "Category": ["category", "metric_category", "kpi_category"],
        "Metric": ["metric", "metric_id", "metric_name"],
        "Value": ["value", "numeric_value", "value_num"],
        "Unit": ["unit", "value_unit"],
        "Source": ["source", "source_name"],
        "Source Tier": ["source_tier", "tier"],
        "Confidence": ["confidence", "confidence_label"],
        "Confidence Score": ["confidence_score", "confidence_num"],
        "Year": ["year", "year_or_range"],
        "Geography": ["geography", "country"],
        "Population": ["population", "segment"],
        "Definition": ["definition", "metric_definition"],
        "Notes": ["notes", "commentary"],
    }

    def pick(row, candidates):
        for c in candidates:
            if c in row.index and pd.notna(row[c]):
                return row[c]
        return ""

    written = 0
    for _, rec in evidence_df.iterrows():
        r = 4 + written
        for col_idx, hdr in enumerate(expected_cols, 1):
            val = pick(rec, colmap[hdr])
            ws.cell(row=r, column=col_idx, value=val)
        written += 1

    # Update subtitle on row 2
    ws.cell(row=2, column=1,
            value=f"{written} records  |  Refreshed: {datetime.now():%Y-%m-%d %H:%M}  |  Source: Data Pipeline tool")
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=10, color="808080", italic=True)
    return written


def stamp_pipeline_metadata(wb, manifest: Dict[str, Any]) -> None:
    """Update the Pipeline Metadata sheet with manifest contents."""
    if "Pipeline Metadata" not in wb.sheetnames:
        return
    ws = wb["Pipeline Metadata"]

    # Rows 5-10 are the "LAST RUN" block
    updates = {
        6: ("Last workbook generation", datetime.now().strftime("%Y-%m-%d %H:%M")),
        7: ("Last pipeline refresh", manifest.get("run_timestamp_utc", "")),
        8: ("Last refresh user", manifest.get("run_user", "")),
        9: ("Pipeline version", manifest.get("pipeline_version", "")),
       10: ("Run ID", datetime.now().strftime("run_%Y%m%d_%H%M%S")),
    }
    for row, (label, value) in updates.items():
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)

    # Clear and rewrite the "ROWS PULLED BY SOURCE" block (starts at row 14, each indication gets a bundle)
    # Find the header row (contains "Source" | "Indication" | ...)
    hdr_row = None
    for r in range(12, 18):
        if ws.cell(row=r, column=1).value == "Source":
            hdr_row = r
            break
    if hdr_row is None:
        return

    first_data_row = hdr_row + 1
    last_data_row = ws.max_row
    # Wipe rows from first_data_row to end, then rewrite
    for r in range(first_data_row, last_data_row + 1):
        for c in range(1, 6):
            ws.cell(row=r, column=c).value = None

    r = first_data_row
    for ind_entry in manifest.get("indications", []):
        ind = ind_entry.get("indication", "?")
        checksums = ind_entry.get("checksums", {})
        # One row per (source=pipeline output file, indication)
        for source, checksum in checksums.items():
            ws.cell(row=r, column=1, value=source)
            ws.cell(row=r, column=2, value=ind)
            ws.cell(row=r, column=3, value="-")
            ws.cell(row=r, column=4, value=ind_entry.get("record_count", "-"))
            ws.cell(row=r, column=5, value=checksum)
            r += 1


def patch_workbook(workbook_path: Path, manifest: Dict[str, Any],
                   results: List[Dict[str, Any]]) -> None:
    """Top-level: open the v5 workbook, apply all patches, save in-place."""
    workbook_path = Path(workbook_path).resolve()
    has_vba = _workbook_zip_has_vba(workbook_path)
    keep_vba = has_vba

    wb = openpyxl.load_workbook(workbook_path, keep_vba=keep_vba)

    if "Lookup Tables" not in wb.sheetnames:
        raise ValueError(f"Workbook missing 'Lookup Tables' sheet: {workbook_path}")
    lookup = wb["Lookup Tables"]

    for r in results:
        if not r["success"]:
            continue
        ind = r["indication"]
        wb_key = pipeline_to_workbook_key(ind)
        if wb_key is None:
            continue

        # Load the indication's KPI scorecard
        scorecard_path_str = r["paths"].get("kpi_scorecard")
        scorecard_df = _load_kpi_scorecard(Path(scorecard_path_str)) if scorecard_path_str else None

        # Patch Model Inputs row (columns B-E)
        if scorecard_df is not None:
            _ = patch_model_inputs(lookup, wb_key, scorecard_df)

        # Replace Evidence sheet contents
        evidence_path_str = r["paths"].get("evidence")
        if evidence_path_str:
            try:
                evidence_df = pd.read_csv(evidence_path_str, encoding="utf-8")
                replace_evidence_sheet(wb, wb_key, evidence_df)
            except Exception:  # noqa: BLE001
                pass

    # Pipeline Metadata stamp
    stamp_pipeline_metadata(wb, manifest)

    wb.save(workbook_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Patch v5.2 workbook with latest pipeline outputs.")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True,
                        help="run_manifest.json written by run_all_indications.py")
    args = parser.parse_args()

    manifest = json.loads(args.manifest.read_text(encoding="utf-8"))

    output_root = args.manifest.resolve().parent

    # Reconstruct minimal results[] list for patch_workbook (paths match runner._safe_name + file_suffix)
    results = []
    for e in manifest.get("indications", []):
        country = e.get("country") or ""
        sfx = _file_suffix(e["indication"], country if country else None)
        results.append(
            {
                "indication": e["indication"],
                "country": country,
                "success": e["success"],
                "record_count": e["record_count"],
                "paths": {
                    "evidence": str(output_root / f"evidence_by_metric_{sfx}.csv"),
                    "kpi_scorecard": str(output_root / f"kpi_scorecard_{sfx}.csv"),
                },
            }
        )

    patch_workbook(args.workbook, manifest, results)
    print(f"Patched: {args.workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
