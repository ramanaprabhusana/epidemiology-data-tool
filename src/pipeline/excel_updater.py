"""
Excel Model Updater — writes pipeline KPI best-values back to
Epidemiology_Forecast_Model_Client_Hub_v6.xlsx, making it the
automated one-stop forecast solution.

Two write targets in the Excel workbook:
  1. Proportion Inputs (Lookup Tables rows 560+, col A=metric_id, col C=value)
     → Stage/LOT/Age/Biomarker proportions; Forecast sheets read these dynamically
  2. Cascade Lookup section (Lookup Tables rows 214-460, col B=option_label, col C=value)
     → Cascade bucket percentages; flows through Control Panel → Forecast sheets

Usage:
    from src.pipeline.excel_updater import update_excel_from_pipeline
    result = update_excel_from_pipeline(
        kpi_csv_dir=Path("output/"),
        excel_path=Path("...v6.xlsx"),
        dry_run=False,
        indications=["CLL", "Hodgkin", "NonHodgkin", "Gastric", "Ovarian", "Prostate"],
    )
    print(result["summary"])
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import openpyxl

logger = logging.getLogger(__name__)

# ── Metric-ID → Cascade option label mapping ──────────────────
# These cascade entries live in the existing Lookup Tables cascade sections
# (rows ~214-460) keyed by the "Option label (MATCH)" in column B.
# excel_updater finds these rows dynamically via cell-content search.
CASCADE_LABEL_MAP: Dict[str, str] = {
    # CLL cascade
    "cascade_watch_wait_pct":        "Watch & Wait | CLL cascade",
    "cascade_drug_ready_pct":        "Drug-Ready | CLL cascade",
    "cascade_1l_pct":                "Treated (1L) | CLL cascade",
    "cascade_2l_pct":                "2L | CLL cascade",
    "cascade_3l_pct":                "3L | CLL cascade",
    "cascade_4l_pct":                "4L | CLL cascade",
    "cascade_5l_pct":                "5L | CLL cascade",
    # Hodgkin cascade
    "hl_cascade_surveillance_pct":   "Active Surveillance | Hodgkin cascade",
    "hl_cascade_1l_pct":             "1L ABVD / BV-AVD | Hodgkin cascade",
    "hl_cascade_cmt_pct":            "Early-Stage Combined-Modality | Hodgkin cascade",
    "hl_cascade_2l_pct":             "2L Salvage + Auto-HSCT | Hodgkin cascade",
    "hl_cascade_refractory_pct":     "Primary Refractory | Hodgkin cascade",
    "hl_cascade_3lplus_pct":         "3L+ (BV / Nivo / Pembro) | Hodgkin cascade",
    "hl_cascade_palliative_pct":     "Palliative / Survivorship | Hodgkin cascade",
    # NHL cascade
    "nhl_cascade_ww_pct":            "Indolent Watch-and-Wait | NonHodgkin cascade",
    "nhl_cascade_1l_aggr_pct":       "1L Aggressive (R-CHOP-like) | NonHodgkin cascade",
    "nhl_cascade_1l_indolent_pct":   "1L Indolent (BR/R2/Obi) | NonHodgkin cascade",
    "nhl_cascade_1l_mcl_pct":        "1L MCL (R-CHOP/R-DHAP+ASCT) | NonHodgkin cascade",
    "nhl_cascade_rr_pct":            "Relapsed/Refractory | NonHodgkin cascade",
    "nhl_cascade_3lplus_pct":        "3L+ (CAR-T / Bispecifics) | NonHodgkin cascade",
    "nhl_cascade_palliative_pct":    "Palliative | NonHodgkin cascade",
    # Gastric cascade
    "gc_cascade_localized_pct":      "Localized Resectable | Gastric cascade",
    "gc_cascade_locally_adv_pct":    "Locally Advanced | Gastric cascade",
    "gc_cascade_met_1l_pct":         "Metastatic 1L | Gastric cascade",
    "gc_cascade_2l_pct":             "2L | Gastric cascade",
    "gc_cascade_3l_pct":             "3L | Gastric cascade",
    "gc_cascade_bsc_pct":            "BSC / Palliative | Gastric cascade",
    # Ovarian cascade
    "ov_cascade_debulk_1l_pct":      "Post-Debulking 1L | Ovarian cascade",
    "ov_cascade_parp_maint_pct":     "1L PARP Maintenance | Ovarian cascade",
    "ov_cascade_plat_sens_pct":      "Platinum-Sensitive 2L | Ovarian cascade",
    "ov_cascade_plat_res_pct":       "Platinum-Resistant 3L | Ovarian cascade",
    "ov_cascade_4lplus_pct":         "4L+ | Ovarian cascade",
    "ov_cascade_bsc_pct":            "BSC / Palliative | Ovarian cascade",
    # Prostate cascade
    "pc_cascade_surveillance_pct":   "Active Surveillance | Prostate cascade",
    "pc_cascade_definitive_pct":     "Definitive Local Tx | Prostate cascade",
    "pc_cascade_bcr_adt_pct":        "BCR / ADT | Prostate cascade",
    "pc_cascade_mcspc_pct":          "mCSPC | Prostate cascade",
    "pc_cascade_nmcrpc_pct":         "nmCRPC/mCRPC 1L-2L | Prostate cascade",
    "pc_cascade_late_mcrpc_pct":     "Late-line mCRPC | Prostate cascade",
}

# ── Core Epi parameter cell locations (Lookup Tables) ─────────
# Format: (sheet_name, row_or_label, col)
# "label_match" means find by matching column B label; "cell" means direct
CORE_EPI_MAP: Dict[str, Dict] = {
    "incidence_rate": {
        "CLL":        ("Lookup Tables", 5, 2),   # row 5, col B
        "Hodgkin":    ("Lookup Tables", 6, 2),
        "NonHodgkin": ("Lookup Tables", 7, 2),
        "Gastric":    ("Lookup Tables", 8, 2),
        "Ovarian":    ("Lookup Tables", 9, 2),
        "Prostate":   ("Lookup Tables", 10, 2),
    },
    "prevalence_anchor": {
        "CLL":        {"label": "Prevalence Anchor (2022) | CLL"},
        "Hodgkin":    {"label": "Prevalence Anchor (2022) | Hodgkin"},
        "NonHodgkin": {"label": "Prevalence Anchor (2022) | NonHodgkin"},
        "Gastric":    {"label": "Prevalence Anchor (2022) | Gastric"},
        "Ovarian":    {"label": "Prevalence Anchor (2022) | Ovarian"},
        "Prostate":   {"label": "Prevalence Anchor (2022) | Prostate"},
    },
    "mortality_rate_aapc": {
        "CLL":        {"label": "Mortality Trend (APC) | CLL"},
        "Hodgkin":    {"label": "Mortality Trend (APC) | Hodgkin"},
        "NonHodgkin": {"label": "Mortality Trend (APC) | NonHodgkin"},
        "Gastric":    {"label": "Mortality Trend (APC) | Gastric"},
        "Ovarian":    {"label": "Mortality Trend (APC) | Ovarian"},
        "Prostate":   {"label": "Mortality Trend (APC) | Prostate"},
    },
}


def _find_rows_by_col_b(ws, search_str: str, max_row: int = 700) -> list[int]:
    """Return all row numbers where column B == search_str."""
    hits = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=2, max_col=2, values_only=True):
        if row[0] == search_str:
            hits.append(ws._current_row if hasattr(ws, "_current_row") else None)
    # Fallback: direct iteration
    if not hits:
        for r in range(1, max_row + 1):
            if ws.cell(row=r, column=2).value == search_str:
                hits.append(r)
    return hits


def _find_row_by_col_a(ws, search_str: str, max_row: int = 700) -> Optional[int]:
    """Return first row number where column A == search_str."""
    for r in range(1, max_row + 1):
        if ws.cell(row=r, column=1).value == search_str:
            return r
    return None


def update_excel_from_pipeline(
    kpi_csv_dir: Path,
    excel_path: Path,
    indications: Optional[List[str]] = None,
    dry_run: bool = True,
    evidence_import_sheet: str = "Evidence Import",
) -> Dict[str, Any]:
    """
    Main entry point: reads KPI scorecard CSVs from kpi_csv_dir,
    writes best_value for each metric to the Excel v6 workbook.

    Args:
        kpi_csv_dir: Directory containing kpi_scorecard_<indication>.csv files
        excel_path: Path to Epidemiology_Forecast_Model_Client_Hub_v6.xlsx
        indications: List of indications to process (default: all 6)
        dry_run: If True, log changes but don't save
        evidence_import_sheet: Sheet name for the Evidence Import status update

    Returns:
        Dict with keys: success, changes, skipped, summary, errors
    """
    if indications is None:
        indications = ["CLL", "Hodgkin", "NonHodgkin", "Gastric", "Ovarian", "Prostate"]

    result: Dict[str, Any] = {
        "success": False,
        "changes": [],
        "skipped": [],
        "errors": [],
        "summary": "",
    }

    if not excel_path.exists():
        result["errors"].append(f"Excel not found: {excel_path}")
        return result

    # ── Config dir for required_metrics YAMLs (excel_metric_id lookup) ──
    config_dir = kpi_csv_dir.parent / "config"
    _ind_to_yaml_suffix = {
        "CLL": "cll", "Hodgkin": "hodgkin", "NonHodgkin": "nhl",
        "Gastric": "gc", "Ovarian": "ovarian", "Prostate": "prostate",
    }
    # Build excel_metric_id → metric_id reverse map per indication
    _excel_id_to_metric: Dict[str, Dict[str, str]] = {}  # ind → {excel_metric_id: metric_id}
    try:
        import yaml as _yaml
        for ind, suffix in _ind_to_yaml_suffix.items():
            yaml_path = config_dir / f"required_metrics_{suffix}.yaml"
            if yaml_path.exists():
                with open(yaml_path) as _f:
                    _data = _yaml.safe_load(_f)
                mapping: Dict[str, str] = {}
                for _m in _data.get("metrics", []):
                    _mid = _m.get("id", "")
                    _eid = _m.get("excel_metric_id", "")
                    if _eid:
                        mapping[_eid] = _mid
                _excel_id_to_metric[ind] = mapping
    except Exception as _e:
        logger.warning(f"Could not load required_metrics YAMLs for excel_metric_id mapping: {_e}")

    # Load all KPI CSVs
    kpi_frames: Dict[str, pd.DataFrame] = {}
    ind_name_map = {
        "CLL": ["CLL", "cll"],
        "Hodgkin": ["Hodgkin", "hodgkin", "Hodgkin_Lymphoma"],
        "NonHodgkin": ["NonHodgkin", "non_hodgkin", "Non_Hodgkin_Lymphoma", "Non-Hodgkin_Lymphoma"],
        "Gastric": ["Gastric", "gastric"],
        "Ovarian": ["Ovarian", "ovarian"],
        "Prostate": ["Prostate", "prostate"],
    }
    for ind in indications:
        for slug in ind_name_map.get(ind, [ind]):
            for pattern in [f"kpi_scorecard_{slug}_US.csv", f"kpi_scorecard_{slug}*.csv", f"kpi_{slug}*.csv"]:
                matches = list(kpi_csv_dir.glob(pattern))
                if matches:
                    try:
                        df = pd.read_csv(matches[0])
                        kpi_frames[ind] = df
                        logger.debug(f"Loaded KPI for {ind}: {matches[0].name}")
                    except Exception as e:
                        result["errors"].append(f"Failed to read {matches[0]}: {e}")
                    break
            if ind in kpi_frames:
                break
        if ind not in kpi_frames:
            result["skipped"].append(f"{ind}: no KPI CSV found in {kpi_csv_dir}")

    if not kpi_frames:
        result["errors"].append("No KPI CSV files found. Run pipeline first.")
        return result

    wb = openpyxl.load_workbook(str(excel_path))
    ws_lt = wb["Lookup Tables"] if "Lookup Tables" in wb.sheetnames else None
    if ws_lt is None:
        result["errors"].append("'Lookup Tables' sheet not found in workbook.")
        wb.close()
        return result

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    changes = []

    # ── Helper ─────────────────────────────────────────────────
    def _apply_value(sheet_name: str, row: int, col: int, metric_id: str,
                     new_val: Any, source: str, confidence: str) -> None:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
        if ws is None:
            return
        cell = ws.cell(row=row, column=col)
        old_val = cell.value
        if dry_run:
            changes.append({
                "metric_id": metric_id, "sheet": sheet_name,
                "cell": f"{chr(64+col)}{row}", "old": old_val, "new": new_val,
                "source": source, "confidence": confidence, "status": "DRY_RUN",
            })
        else:
            cell.value = new_val
            changes.append({
                "metric_id": metric_id, "sheet": sheet_name,
                "cell": f"{chr(64+col)}{row}", "old": old_val, "new": new_val,
                "source": source, "confidence": confidence, "status": "UPDATED",
            })

    # ── 1) Proportion Inputs section (Lookup Tables col A=excel_metric_id, col C=value) ─
    for ind, kpi_df in kpi_frames.items():
        if kpi_df.empty:
            continue
        if "metric_id" not in kpi_df.columns or "best_value" not in kpi_df.columns:
            result["errors"].append(f"{ind}: KPI CSV missing metric_id or best_value columns")
            continue
        kpi_map = dict(zip(kpi_df["metric_id"], kpi_df["best_value"]))
        src_map = dict(zip(kpi_df["metric_id"], kpi_df.get("best_source", pd.Series())))
        conf_map = dict(zip(kpi_df["metric_id"], kpi_df.get("validation_status", pd.Series())))

        # Build reverse lookup: excel_metric_id → best_value (for Proportion Inputs column A)
        excel_id_map = _excel_id_to_metric.get(ind, {})  # excel_metric_id → metric_id
        kpi_by_excel_id: Dict[str, Any] = {}
        for excel_mid, kpi_mid in excel_id_map.items():
            if kpi_mid in kpi_map and not (
                pd.isna(kpi_map[kpi_mid]) if hasattr(pd, "isna") else kpi_map[kpi_mid] != kpi_map[kpi_mid]
            ) and kpi_map[kpi_mid] != "":
                kpi_by_excel_id[excel_mid] = (kpi_map[kpi_mid], src_map.get(kpi_mid, "Pipeline"),
                                               conf_map.get(kpi_mid, ""))

        # Search Proportion Inputs section (rows 560-700, col A = excel_metric_id, col C = value)
        for r in range(560, 710):
            mid_cell = ws_lt.cell(row=r, column=1).value
            if mid_cell is None:
                continue
            mid_cell = str(mid_cell).strip()

            # Try match via excel_metric_id (primary path)
            if mid_cell in kpi_by_excel_id:
                new_val, src, conf = kpi_by_excel_id[mid_cell]
            elif mid_cell in kpi_map:
                # Fallback: direct metric_id match (handles plain metric IDs if present)
                new_val = kpi_map[mid_cell]
                if pd.isna(new_val) if hasattr(pd, "isna") else (new_val != new_val):
                    continue
                src = str(src_map.get(mid_cell, "Pipeline")).strip()
                conf = str(conf_map.get(mid_cell, "")).strip()
            else:
                continue

            try:
                float_val = float(new_val)
            except (ValueError, TypeError):
                continue
            # Normalize: if existing cell is in decimal form (0-1) and new value is pct (>1),
            # convert to decimal. Skip values that are clearly absolute counts (>150 after /100).
            existing_val = ws_lt.cell(row=r, column=3).value
            if isinstance(existing_val, (int, float)) and 0 <= existing_val <= 1.0:
                if float_val > 1.0:
                    float_val = float_val / 100.0
                if float_val > 1.5:  # still > 100% after normalisation → skip (likely count)
                    logger.debug(f"Skipping {mid_cell} row {r}: value {float_val:.2f} looks like a count")
                    continue
            _apply_value("Lookup Tables", r, 3, mid_cell, round(float_val, 4), src, conf)
            # Also update source and last_updated columns
            if not dry_run:
                ws_lt.cell(row=r, column=4).value = src
                ws_lt.cell(row=r, column=5).value = now_str

        # ── 2) Cascade Lookup sections ───────────────────────
        for metric_id, option_label in CASCADE_LABEL_MAP.items():
            if metric_id not in kpi_map:
                continue
            new_val = kpi_map[metric_id]
            if pd.isna(new_val) if hasattr(pd, "isna") else (new_val != new_val):
                continue
            # Find row in Lookup Tables where col B == option_label
            for r in range(200, 470):
                if ws_lt.cell(row=r, column=2).value == option_label:
                    src = str(src_map.get(metric_id, "Pipeline")).strip()
                    conf = str(conf_map.get(metric_id, "")).strip()
                    _apply_value("Lookup Tables", r, 3, metric_id,
                                 float(new_val), src, conf)
                    if not dry_run:
                        ws_lt.cell(row=r, column=4).value = src
                    break

        # ── 3) Core Epi parameters (incidence_rate, prevalence_anchor, etc.) ─
        for epi_metric, ind_map_entry in CORE_EPI_MAP.items():
            if epi_metric not in kpi_map:
                continue
            new_val = kpi_map[epi_metric]
            if pd.isna(new_val) if hasattr(pd, "isna") else (new_val != new_val):
                continue
            target = ind_map_entry.get(ind)
            if target is None:
                continue
            if isinstance(target, tuple):  # (sheet, row, col)
                sheet_name, row, col = target
                src = str(src_map.get(epi_metric, "Pipeline"))
                conf = str(conf_map.get(epi_metric, ""))
                _apply_value(sheet_name, row, col, epi_metric, float(new_val), src, conf)
            elif isinstance(target, dict) and "label" in target:  # find by label
                for r in range(200, 470):
                    if ws_lt.cell(row=r, column=2).value == target["label"]:
                        src = str(src_map.get(epi_metric, "Pipeline"))
                        conf = str(conf_map.get(epi_metric, ""))
                        _apply_value("Lookup Tables", r, 3, epi_metric, float(new_val), src, conf)
                        break

    # ── 4) Update Evidence Import sheet pipeline columns (F, G, H, I) ──
    ws_ei = wb[evidence_import_sheet] if evidence_import_sheet in wb.sheetnames else None
    if ws_ei is not None:
        for r in range(4, 300):
            ind_cell = ws_ei.cell(row=r, column=1).value
            metric_cell = ws_ei.cell(row=r, column=3).value
            mid_hint = ws_ei.cell(row=r, column=10).value  # Status col has metric_id hint? No
            if ind_cell not in (kpi_frames.keys() | {None}):
                continue
            # Find matching KPI row by indication + metric label fuzzy match
            if ind_cell and ind_cell in kpi_frames and metric_cell:
                kpi_df = kpi_frames[ind_cell]
                if "label" in kpi_df.columns:
                    match = kpi_df[kpi_df["label"].str.lower() == str(metric_cell).lower()]
                    if not match.empty and not dry_run:
                        best_val = match.iloc[0].get("best_value")
                        best_src = match.iloc[0].get("best_source", "")
                        val_status = match.iloc[0].get("validation_status", "")
                        if best_val is not None and not (isinstance(best_val, float) and pd.isna(best_val)):
                            ws_ei.cell(row=r, column=6).value = float(best_val)
                            ws_ei.cell(row=r, column=7).value = str(best_src)
                            ws_ei.cell(row=r, column=8).value = str(val_status)
                            ws_ei.cell(row=r, column=9).value = now_str

    # ── Save ───────────────────────────────────────────────────
    if not dry_run and changes:
        try:
            wb.save(str(excel_path))
            logger.info(f"Saved Excel v6: {excel_path}")
        except Exception as e:
            result["errors"].append(f"Failed to save Excel: {e}")
            result["summary"] = f"Error saving: {e}"
            return result

    n_updated = sum(1 for c in changes if c.get("status") == "UPDATED")
    n_dry = sum(1 for c in changes if c.get("status") == "DRY_RUN")
    result["success"] = True
    result["changes"] = changes
    result["summary"] = (
        f"{'[DRY RUN] ' if dry_run else ''}"
        f"{n_updated + n_dry} metric values "
        f"{'would be ' if dry_run else ''}written to Excel v6 "
        f"across {len(kpi_frames)} indications."
    )
    return result


def run_updater_cli():
    """CLI entry point: python -m src.pipeline.excel_updater <kpi_dir> <excel_path> [--apply]"""
    import sys
    import argparse
    parser = argparse.ArgumentParser(description="Update Excel v6 from pipeline KPI outputs")
    parser.add_argument("kpi_dir", help="Directory containing kpi_scorecard_*.csv files")
    parser.add_argument("excel_path", help="Path to Epidemiology_Forecast_Model_Client_Hub_v6.xlsx")
    parser.add_argument("--apply", action="store_true", help="Apply changes (default: dry run)")
    parser.add_argument("--indication", nargs="+", help="Specific indications to process")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    result = update_excel_from_pipeline(
        kpi_csv_dir=Path(args.kpi_dir),
        excel_path=Path(args.excel_path),
        indications=args.indication,
        dry_run=not args.apply,
    )
    print(result["summary"])
    if result["errors"]:
        print("ERRORS:")
        for e in result["errors"]:
            print(f"  {e}")
    if result["skipped"]:
        print("SKIPPED:")
        for s in result["skipped"]:
            print(f"  {s}")
    if args.apply:
        for ch in result["changes"][:20]:
            print(f"  {ch['status']}: {ch['metric_id']} → {ch['sheet']}!{ch['cell']} = {ch['new']}")
    sys.exit(0 if result["success"] else 1)


if __name__ == "__main__":
    run_updater_cli()
